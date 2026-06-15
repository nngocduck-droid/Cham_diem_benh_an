from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from typing import List
import google.generativeai as genai
import fitz  # PyMuPDF
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
import tempfile
from datetime import datetime
from difflib import SequenceMatcher
import pathlib

load_dotenv()

genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
model = genai.GenerativeModel("gemini-2.5-flash")

app = FastAPI(title="Chấm Bệnh Án Y Khoa")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Xác thực đơn giản ──
APP_PASSWORD = os.getenv("APP_PASSWORD", "bacsi2024")

@app.post("/login")
def login(data: dict):
    if data.get("password") == APP_PASSWORD:
        return {"success": True, "message": "Đăng nhập thành công"}
    raise HTTPException(status_code=401, detail="Mật khẩu không đúng")

def check_auth(password: str):
    if password != APP_PASSWORD:
        raise HTTPException(status_code=401, detail="Không có quyền truy cập")

# History file path - lưu cùng thư mục backend
HISTORY_FILE = pathlib.Path(__file__).parent / "lich_su_cham.json"

RUBRIC = """
Bạn là giảng viên y khoa chấm bệnh án sinh viên. Hãy chấm điểm theo rubric sau (thang 10 điểm, điểm lẻ 0.25):

=== PHẦN A: CÁC ĐỀ MỤC (tối đa 7.0 điểm) ===

1. Hành chính (tối đa 0.25đ): Đủ 9 mục: Họ tên, Tuổi, Giới, Nghề nghiệp, Địa chỉ, Ngày vào viện, Ngày làm bệnh án, Giường/Phòng, Mã bệnh nhân.
   - 0.25: Đủ 9 mục | 0: Thiếu từ 1 mục trở lên

2. Lý do vào viện (tối đa 0.25đ): Nêu đúng và đủ lý do chính vào viện.
   - 0.25: Đúng, rõ ràng | 0: Sai, thiếu hoặc không có

3. Bệnh sử (tối đa 0.75đ): Mô tả đủ thời gian, hoàn cảnh, đặc điểm, diễn biến, triệu chứng kèm theo, điều trị trước đó, triệu chứng hiện tại.
   - 0.75: Đủ tất cả | 0.5: Thiếu 1 yếu tố | 0.25: Thiếu 2-3 yếu tố | 0: Qua loa hoặc không có

4. Tiền sử bệnh (tối đa 0.25đ): Hút thuốc (số bao-năm), bệnh mạn tính, dị ứng, dịch tễ, tiền sử gia đình.
   - 0.25: Đủ ≥3 mục | 0: Dưới 3 mục hoặc không có

5. Khám bệnh (tối đa 1.0đ): Khám vào viện + Khám hiện tại đủ: toàn thân, hô hấp (nhìn-sờ-gõ-nghe), tim mạch, tiêu hoá, thận-tiết niệu, cơ xương khớp, cơ quan khác.
   - 1.0: Đủ chi tiết | 0.75: Thiếu 1-2 hệ | 0.5: Thiếu ≥3 hệ | 0.25: Sơ sài | 0: Không có

6. Tóm tắt bệnh án (tối đa 0.75đ): Đúng chuẩn: giới-tuổi-lý do-số ngày bệnh, hội chứng dương tính, âm tính, tiền sử.
   - 0.75: Đủ cấu trúc | 0.5: Thiếu âm tính/tiền sử | 0.25: Chưa đầy đủ | 0: Rất sơ sài hoặc không có

7. Chẩn đoán sơ bộ (tối đa 0.75đ): Phù hợp với bệnh sử và tóm tắt bệnh án.
   - 0.75: Đúng, có căn cứ | 0.5: Đúng hướng, chưa đầy đủ | 0.25: Mơ hồ | 0: Sai

8. Đề xuất cận lâm sàng (tối đa 0.75đ): Đề xuất đúng và nêu lý do cho từng xét nghiệm.
   - 0.75: Đầy đủ, có lý do | 0.5: Đủ nhưng thiếu một số lý do | 0.25: Thiếu xét nghiệm quan trọng hoặc không có lý do | 0: Sai/không có

9. Kết quả cận lâm sàng (tối đa 0.75đ): Trình bày đúng thứ tự, có nhận định bình thường/bất thường.
   - 0.75: Đầy đủ, nhận định chính xác | 0.5: Đủ, nhận định chưa đầy đủ | 0.25: Thiếu nhiều hoặc thiếu nhận định | 0: Không có

10. Chẩn đoán xác định (tối đa 0.75đ): Đúng và có biện luận dựa trên lâm sàng + cận lâm sàng.
    - 0.75: Đúng, biện luận đầy đủ | 0.5: Đúng, biện luận chưa đủ | 0.25: Sai một phần hoặc thiếu biện luận | 0: Sai

11. Điều trị (tối đa 0.75đ): Nguyên tắc điều trị, kế hoạch (thuốc + không thuốc + phác đồ), tư vấn sau ra viện.
    - 0.75: Đầy đủ 3 phần | 0.5: Đủ nguyên tắc và phác đồ, thiếu tư vấn | 0.25: Chỉ nêu được một phần | 0: Không có

=== PHẦN B: TƯ DUY LÂM SÀNG (tối đa 3.0 điểm) ===

12. Tư duy lâm sàng mạch lạc (tối đa 3.0đ): Đánh giá tính thống nhất và logic trong toàn bộ bệnh án: bệnh sử → khám bệnh → tóm tắt → chẩn đoán sơ bộ → cận lâm sàng → chẩn đoán xác định → điều trị. Các phần có ăn khớp, bổ sung cho nhau không? Có mâu thuẫn nội tại không? Biện luận có được xây dựng nhất quán từ dữ liệu lâm sàng không?
    - 3.0: Tư duy rõ ràng, mạch lạc, nhất quán xuyên suốt; biện luận logic từ triệu chứng → chẩn đoán → điều trị
    - 2.25: Tư duy nhất quán ở hầu hết các phần, có vài điểm chưa ăn khớp nhưng không ảnh hưởng kết luận
    - 1.5: Có sự thống nhất cơ bản nhưng xuất hiện mâu thuẫn ở một số phần quan trọng
    - 0.75: Tư duy rời rạc, các phần thiếu liên kết, mâu thuẫn rõ ràng
    - 0: Không có tư duy lâm sàng rõ ràng, các phần không ăn khớp hoặc mâu thuẫn nhau

Ngoài chấm điểm, hãy trích xuất thông tin hành chính của bệnh nhân và liệt kê cụ thể các phần sinh viên cần sửa.

Hãy trả về JSON với định dạng chính xác sau (không thêm bất kỳ văn bản nào khác):
{
  "tong_diem": <số thực, tổng của tất cả 12 mục, tối đa 10.0>,
  "xep_loai": "<Xuất sắc/Giỏi/Khá/Trung bình/Yếu>",
  "benh_nhan": {
    "ho_ten": "<họ tên bệnh nhân hoặc viết tắt nếu có>",
    "tuoi": "<tuổi>",
    "gioi": "<giới tính>",
    "ma_benh_nhan": "<mã bệnh nhân nếu có>"
  },
  "chi_tiet": [
    {
      "stt": 1,
      "ten_muc": "Hành chính",
      "diem_toi_da": 0.25,
      "diem_dat": <số>,
      "nhan_xet": "<nhận xét ngắn gọn>"
    }
  ],
  "cac_phan_can_sua": [
    "<mô tả cụ thể phần cần sửa 1>",
    "<mô tả cụ thể phần cần sửa 2>"
  ],
  "nhan_xet_chung": "<nhận xét tổng thể về bệnh án>"
}
"""

def extract_text_from_pdf(file_bytes: bytes) -> str:
    with fitz.open(stream=file_bytes, filetype="pdf") as doc:
        text = ""
        for page in doc:
            text += page.get_text()
    return text.strip()

def load_history() -> list:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

def save_history(records: list):
    HISTORY_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def check_plagiarism(records: list) -> list:
    """So sánh tất cả các cặp bệnh án để phát hiện trùng lặp"""
    pairs = []
    n = len(records)
    for i in range(n):
        for j in range(i + 1, n):
            a = records[i]
            b = records[j]

            # So sánh thông tin bệnh nhân
            bn_a = a.get("benh_nhan", {})
            bn_b = b.get("benh_nhan", {})

            same_patient = False
            patient_clues = []

            # Kiểm tra tên
            ten_a = bn_a.get("ho_ten", "")
            ten_b = bn_b.get("ho_ten", "")
            if ten_a and ten_b and similarity(ten_a, ten_b) > 0.8:
                same_patient = True
                patient_clues.append(f"Tên bệnh nhân giống nhau: '{ten_a}' – '{ten_b}'")

            # Kiểm tra mã bệnh nhân
            ma_a = bn_a.get("ma_benh_nhan", "")
            ma_b = bn_b.get("ma_benh_nhan", "")
            if ma_a and ma_b and ma_a.strip() == ma_b.strip():
                same_patient = True
                patient_clues.append(f"Mã bệnh nhân trùng nhau: '{ma_a}'")

            # Kiểm tra tuổi + giới
            if (bn_a.get("tuoi") == bn_b.get("tuoi") and
                bn_a.get("gioi", "").lower() == bn_b.get("gioi", "").lower() and
                bn_a.get("tuoi")):
                patient_clues.append(f"Cùng tuổi ({bn_a.get('tuoi')}) và giới tính ({bn_a.get('gioi')})")

            # So sánh nội dung bệnh án
            nxA = a.get("nhan_xet_chung", "")
            nxB = b.get("nhan_xet_chung", "")
            content_sim = similarity(nxA, nxB) if nxA and nxB else 0

            # Chỉ báo cáo nếu có dấu hiệu trùng lặp
            if same_patient or content_sim > 0.7:
                muc_do = "🔴 Rất cao" if content_sim > 0.85 or (same_patient and content_sim > 0.6) else \
                         "🟠 Cao" if content_sim > 0.7 or same_patient else "🟡 Trung bình"
                pairs.append({
                    "sv1": a.get("ten_file", ""),
                    "sv2": b.get("ten_file", ""),
                    "muc_do": muc_do,
                    "do_tuong_dong": round(content_sim * 100, 1),
                    "chi_tiet": patient_clues if patient_clues else ["Nội dung bệnh án tương đồng cao"]
                })
    return pairs

@app.post("/cham-nhieu")
async def cham_nhieu(files: List[UploadFile] = File(...)):
    results = []
    errors = []

    for file in files:
        if not file.filename.endswith(".pdf"):
            errors.append({"file": file.filename, "loi": "Không phải file PDF"})
            continue
        try:
            content = await file.read()
            text = extract_text_from_pdf(content)
            if not text:
                errors.append({"file": file.filename, "loi": "Không đọc được nội dung PDF"})
                continue

            prompt = f"{RUBRIC}\n\nNỘI DUNG BỆNH ÁN:\n{text}"
            response = model.generate_content(prompt)
            raw = response.text.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
            result = json.loads(raw.strip())
            result["ten_file"] = file.filename
            result["thoi_gian"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            results.append(result)
        except Exception as e:
            errors.append({"file": file.filename, "loi": str(e)})

    # Lưu vào lịch sử
    if results:
        history = load_history()
        history.extend(results)
        save_history(history)

    # Kiểm tra đạo văn trong lô này
    plagiarism = check_plagiarism(results) if len(results) > 1 else []

    return {
        "results": results,
        "errors": errors,
        "plagiarism": plagiarism,
        "tong_so": len(results),
        "loi_so": len(errors)
    }

@app.get("/lich-su")
def get_lich_su():
    return load_history()

@app.delete("/lich-su")
def xoa_lich_su():
    save_history([])
    return {"message": "Đã xoá toàn bộ lịch sử"}

@app.post("/kiem-tra-dao-van")
def kiem_tra_dao_van():
    """Kiểm tra đạo văn trong toàn bộ lịch sử"""
    history = load_history()
    if len(history) < 2:
        return {"pairs": [], "message": "Cần ít nhất 2 bệnh án để so sánh"}
    pairs = check_plagiarism(history)
    return {"pairs": pairs, "tong_cap": len(pairs)}

@app.post("/xuat-excel-nhieu")
async def xuat_excel_nhieu(data: dict):
    results = data.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="Không có dữ liệu")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tổng hợp ──
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    blue = PatternFill("solid", fgColor="1F6AA5")
    light = PatternFill("solid", fgColor="D6E8F7")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    white_bold = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11)
    bold = Font(name="Times New Roman", bold=True, size=11)
    normal = Font(name="Times New Roman", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1.merge_cells("A1:H1")
    ws1["A1"] = "BẢNG TỔNG HỢP ĐIỂM CHẤM BỆNH ÁN Y KHOA"
    ws1["A1"].font = Font(name="Times New Roman", bold=True, size=14, color="1F6AA5")
    ws1["A1"].alignment = center

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Tổng số bệnh án: {len(results)}"
    ws1["A2"].font = Font(name="Times New Roman", italic=True, size=10, color="666666")
    ws1["A2"].alignment = center
    ws1.append([])

    headers = ["STT", "File bệnh án", "Họ tên BN", "Tuổi", "Giới", "Tổng điểm", "Xếp loại", "Thời gian"]
    ws1.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=col)
        c.fill = blue; c.font = white_bold; c.alignment = center; c.border = bd

    for i, r in enumerate(results, 1):
        bn = r.get("benh_nhan", {})
        diem = r.get("tong_diem", 0)
        fill = green_fill if diem >= 5 else red_fill
        row = [i, r.get("ten_file",""), bn.get("ho_ten",""), bn.get("tuoi",""),
               bn.get("gioi",""), diem, r.get("xep_loai",""), r.get("thoi_gian","")]
        ws1.append(row)
        ro = ws1.max_row
        for col in range(1, 9):
            c = ws1.cell(row=ro, column=col)
            c.fill = light if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            c.font = normal; c.border = bd
            c.alignment = center if col != 2 else left
        ws1.cell(ro, 6).fill = fill
        ws1.cell(ro, 6).font = Font(name="Times New Roman", bold=True, size=11)

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 8
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 14
    ws1.column_dimensions["H"].width = 18

    # ── Sheet 2: Chi tiết từng bệnh án ──
    for r in results:
        sname = r.get("ten_file","").replace(".pdf","")[:28]
        ws = wb.create_sheet(title=sname)
        ws.merge_cells("A1:F1")
        ws["A1"] = f"KẾT QUẢ CHẤM: {r.get('ten_file','')}"
        ws["A1"].font = Font(name="Times New Roman", bold=True, size=13, color="1F6AA5")
        ws["A1"].alignment = center

        bn = r.get("benh_nhan", {})
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Bệnh nhân: {bn.get('ho_ten','')} | Tuổi: {bn.get('tuoi','')} | Giới: {bn.get('gioi','')} | Mã: {bn.get('ma_benh_nhan','')}"
        ws["A2"].font = Font(name="Times New Roman", italic=True, size=10)
        ws["A2"].alignment = center
        ws.append([])

        hdrs = ["STT", "Nội dung", "Điểm tối đa", "Điểm đạt", "Tỉ lệ", "Nhận xét"]
        ws.append(hdrs)
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col)
            c.fill = blue; c.font = white_bold; c.alignment = center; c.border = bd

        for idx, item in enumerate(r.get("chi_tiet", []), 1):
            pct = f"{round(item['diem_dat']/item['diem_toi_da']*100)}%" if item['diem_toi_da'] > 0 else "0%"
            ws.append([item["stt"], item["ten_muc"], item["diem_toi_da"], item["diem_dat"], pct, item["nhan_xet"]])
            ro = ws.max_row
            fill_row = light if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col in range(1, 7):
                c = ws.cell(row=ro, column=col)
                c.fill = fill_row; c.font = normal; c.border = bd
                c.alignment = center if col in [1,3,4,5] else left

        # Total
        tr = ws.max_row + 1
        ws.merge_cells(f"A{tr}:B{tr}")
        ws.cell(tr,1,"TỔNG ĐIỂM").font = white_bold
        ws.cell(tr,1).fill = blue; ws.cell(tr,1).alignment = center; ws.cell(tr,1).border = bd
        ws.cell(tr,3,10.0).fill = blue; ws.cell(tr,3).font = white_bold; ws.cell(tr,3).alignment = center; ws.cell(tr,3).border = bd
        ws.cell(tr,4, r.get("tong_diem",0)).font = Font(name="Times New Roman", bold=True, size=13, color="C00000")
        ws.cell(tr,4).fill = green_fill; ws.cell(tr,4).alignment = center; ws.cell(tr,4).border = bd
        ws.cell(tr,5, r.get("xep_loai","")).font = bold; ws.cell(tr,5).fill = green_fill; ws.cell(tr,5).alignment = center; ws.cell(tr,5).border = bd
        ws.cell(tr,6, r.get("nhan_xet_chung","")).font = normal; ws.cell(tr,6).fill = green_fill; ws.cell(tr,6).alignment = left; ws.cell(tr,6).border = bd

        # Các phần cần sửa
        phan_sua = r.get("cac_phan_can_sua", [])
        if phan_sua:
            sr = ws.max_row + 1
            ws.merge_cells(f"A{sr}:B{sr}")
            ws.cell(sr, 1, "CÁC PHẦN CẦN SỬA").font = Font(name="Times New Roman", bold=True, size=11, color="C00000")
            ws.cell(sr, 1).fill = PatternFill("solid", fgColor="FCE4D6"); ws.cell(sr, 1).alignment = center; ws.cell(sr, 1).border = bd
            ws.merge_cells(f"C{sr}:F{sr}")
            sua_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(phan_sua))
            ws.cell(sr, 3, sua_text).font = normal; ws.cell(sr, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); ws.cell(sr, 3).border = bd
            ws.cell(sr, 3).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.row_dimensions[sr].height = max(15 * len(phan_sua), 30)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 13
        ws.column_dimensions["D"].width = 11
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 55

    # ── Sheet 3: Kiểm tra đạo văn ──
    plagiarism = data.get("plagiarism", [])
    ws3 = wb.create_sheet(title="Kiểm tra đạo văn")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "DANH SÁCH CÁC CẶP BỆNH ÁN NGHI VẤN TRÙNG LẶP"
    ws3["A1"].font = Font(name="Times New Roman", bold=True, size=13, color="C00000")
    ws3["A1"].alignment = center
    ws3.append([])

    if plagiarism:
        hdrs3 = ["Sinh viên 1", "Sinh viên 2", "Mức độ", "Độ tương đồng (%)", "Chi tiết"]
        ws3.append(hdrs3)
        for col, h in enumerate(hdrs3, 1):
            c = ws3.cell(row=3, column=col)
            c.fill = PatternFill("solid", fgColor="C00000")
            c.font = white_bold; c.alignment = center; c.border = bd
        for p in plagiarism:
            ws3.append([p["sv1"], p["sv2"], p["muc_do"], p["do_tuong_dong"],
                        " | ".join(p.get("chi_tiet", []))])
            ro = ws3.max_row
            for col in range(1, 6):
                c = ws3.cell(row=ro, column=col)
                c.fill = PatternFill("solid", fgColor="FFF2CC")
                c.font = normal; c.border = bd
                c.alignment = center if col in [3,4] else left
    else:
        ws3.merge_cells("A3:E3")
        ws3["A3"] = "✅ Không phát hiện trùng lặp đáng kể giữa các bệnh án"
        ws3["A3"].font = Font(name="Times New Roman", size=12, color="2D7A4F")
        ws3["A3"].alignment = center

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 50

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    fname = f"KetQua_BenhAn_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    return FileResponse(tmp.name, filename=fname,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

STATIC_DIR = pathlib.Path(__file__).parent / "static"

@app.get("/", response_class=HTMLResponse)
def serve_frontend():
    index = STATIC_DIR / "index.html"
    if index.exists():
        return index.read_text(encoding="utf-8")
    return HTMLResponse("<h1>Frontend not found</h1>", status_code=404)

@app.get("/health")
def health():
    return {"status": "ok"}
